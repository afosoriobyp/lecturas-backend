from __future__ import annotations

import csv
import io
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.api.dependencies import SessionDep
from app.core.exceptions import NotFoundException
from app.core.security import get_current_user, require_roles
from app.models.meter import Meter
from app.models.ruta import Ruta
from app.models.ruta_asignada import RutaAsignada
from app.models.user import User
from app.schemas.admin import LectorCreate, LectorOut, LectorResetPassword, LectorUpdate
from app.schemas.ruta_asignada import AsignacionCreate, AsignacionOut, AsignacionUpdate
from app.services.audit import log_action

router = APIRouter(
    prefix="/admin",
    tags=["Admin - Lectores y Asignaciones"],
)


@router.get("/lectores", response_model=list[LectorOut])
async def list_lectores(
    session: SessionDep,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    result = await session.execute(
        select(User).where(User.rol == "lector").order_by(User.username)
    )
    return list(result.scalars().all())


@router.post("/lectores", response_model=LectorOut, status_code=201)
async def create_lector(
    session: SessionDep,
    payload: LectorCreate,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    from app.services.auth import register_user

    user = await register_user(
        db=session,
        username=payload.username,
        password=payload.password,
        email=payload.email,
        full_name=payload.full_name,
        rol="lector",
    )

    user.id_tercero = payload.id_tercero
    await session.flush()

    await log_action(
        session, accion="lector.create", entidad_tipo="user",
        entidad_id=user.id, actor_id=str(current_user.id),
        detalle={"username": user.username, "id_tercero": user.id_tercero},
    )

    return user


@router.patch("/lectores/{lector_id}", response_model=LectorOut)
async def update_lector(
    session: SessionDep,
    lector_id: UUID,
    payload: LectorUpdate,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    lector = await session.get(User, lector_id)
    if not lector or lector.rol != "lector":
        raise NotFoundException("Lector")

    update_data = payload.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(lector, key, value)

    await session.flush()

    await log_action(
        session, accion="lector.update", entidad_tipo="user",
        entidad_id=lector.id, actor_id=str(current_user.id),
        detalle={"updated": list(update_data.keys()), "username": lector.username},
    )

    return lector


@router.delete("/lectores/{lector_id}", status_code=204)
async def delete_lector(
    session: SessionDep,
    lector_id: UUID,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    lector = await session.get(User, lector_id)
    if not lector or lector.rol != "lector":
        raise NotFoundException("Lector")

    await session.delete(lector)

    await log_action(
        session, accion="lector.delete", entidad_tipo="user",
        entidad_id=lector_id, actor_id=str(current_user.id),
        detalle={"username": lector.username},
    )


@router.post("/lectores/{lector_id}/reset-password", response_model=LectorOut)
async def reset_lector_password(
    session: SessionDep,
    lector_id: UUID,
    payload: LectorResetPassword,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    from app.core.security import hash_password

    lector = await session.get(User, lector_id)
    if not lector or lector.rol != "lector":
        raise NotFoundException("Lector")

    lector.hashed_password = hash_password(payload.new_password)
    await session.flush()

    await log_action(
        session, accion="lector.reset_password", entidad_tipo="user",
        entidad_id=lector.id, actor_id=str(current_user.id),
        detalle={"username": lector.username},
    )

    return lector


@router.post("/asignaciones", response_model=AsignacionOut, status_code=201)
async def create_asignacion(
    session: SessionDep,
    payload: AsignacionCreate,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    lector = await session.get(User, payload.lector_id)
    if not lector or lector.rol != "lector":
        raise NotFoundException("Lector")

    ruta = await session.get(Ruta, payload.ruta_id)
    if not ruta:
        raise NotFoundException("Ruta")

    existing = await session.execute(
        select(RutaAsignada).where(
            RutaAsignada.lector_id == payload.lector_id,
            RutaAsignada.ruta_id == payload.ruta_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="La ruta ya está asignada a este lector",
        )

    asignacion = RutaAsignada(
        lector_id=payload.lector_id,
        ruta_id=payload.ruta_id,
    )
    session.add(asignacion)
    await session.flush()

    await log_action(
        session, accion="asignacion.create", entidad_tipo="ruta_asignada",
        entidad_id=asignacion.id, actor_id=str(current_user.id),
        detalle={"lector_id": str(payload.lector_id), "ruta_id": str(payload.ruta_id)},
    )

    return AsignacionOut(
        id=asignacion.id,
        lector_id=asignacion.lector_id,
        ruta_id=asignacion.ruta_id,
        ruta_codigo=ruta.codigo,
        ruta_nombre=ruta.nombre,
        ruta_zona=ruta.zona,
    )


@router.put("/asignaciones", response_model=list[AsignacionOut])
async def replace_asignaciones(
    session: SessionDep,
    payload: AsignacionUpdate,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    lector = await session.get(User, payload.lector_id)
    if not lector or lector.rol != "lector":
        raise NotFoundException("Lector")

    for ruta_id in payload.ruta_ids:
        ruta = await session.get(Ruta, ruta_id)
        if not ruta:
            raise NotFoundException(f"Ruta {ruta_id}")

    old = await session.execute(
        select(RutaAsignada).where(RutaAsignada.lector_id == payload.lector_id)
    )
    for row in old.scalars().all():
        await session.delete(row)

    nuevas: list[RutaAsignada] = []
    for ruta_id in payload.ruta_ids:
        a = RutaAsignada(lector_id=payload.lector_id, ruta_id=ruta_id)
        session.add(a)
        nuevas.append(a)
    await session.flush()

    rutas_map = {
        r.id: r
        for r in (
            await session.execute(
                select(Ruta).where(Ruta.id.in_(payload.ruta_ids))
            )
        ).scalars().all()
    }

    await log_action(
        session, accion="asignacion.replace", entidad_tipo="ruta_asignada",
        entidad_id=payload.lector_id, actor_id=str(current_user.id),
        detalle={"lector_id": str(payload.lector_id), "ruta_ids": [str(rid) for rid in payload.ruta_ids]},
    )

    return [
        AsignacionOut(
            id=a.id,
            lector_id=a.lector_id,
            ruta_id=a.ruta_id,
            ruta_codigo=rutas_map[a.ruta_id].codigo,
            ruta_nombre=rutas_map[a.ruta_id].nombre,
            ruta_zona=rutas_map[a.ruta_id].zona,
        )
        for a in nuevas
    ]


@router.get("/asignaciones/{lector_id}", response_model=list[AsignacionOut])
async def get_asignaciones(
    session: SessionDep,
    lector_id: UUID,
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    result = await session.execute(
        select(RutaAsignada)
        .options(joinedload(RutaAsignada.ruta))
        .where(RutaAsignada.lector_id == lector_id)
    )
    asignaciones = result.unique().scalars().all()
    if not asignaciones:
        lector = await session.get(User, lector_id)
        if not lector:
            raise NotFoundException("Lector")

    return [
        AsignacionOut(
            id=a.id,
            lector_id=a.lector_id,
            ruta_id=a.ruta_id,
            ruta_codigo=a.ruta.codigo,
            ruta_nombre=a.ruta.nombre,
            ruta_zona=a.ruta.zona,
        )
        for a in asignaciones
    ]


@router.post("/cargar-medidores")
async def cargar_medidores(
    session: SessionDep,
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    filename = file.filename or ""
    content = await file.read()

    rows: list[dict[str, str]] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="El archivo XLSX no tiene hojas")
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    else:
        raise HTTPException(
            status_code=400,
            detail="Formato no soportado. Use CSV o XLSX",
        )

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    required_fields = {"codigo_medidor"}
    creados = 0
    errores: list[dict] = []

    for i, row in enumerate(rows, start=2):
        codigo = row.get("codigo_medidor", "").strip()
        if not codigo:
            errores.append({"fila": i, "error": "codigo_medidor es requerido"})
            continue

        existing = await session.execute(
            select(Meter).where(Meter.codigo_medidor == codigo)
        )
        if existing.scalar_one_or_none():
            errores.append({"fila": i, "codigo": codigo, "error": "Ya existe"})
            continue

        lector_id: UUID | None = None
        lector_username = row.get("lector_username", "").strip()
        if lector_username:
            lector_result = await session.execute(
                select(User).where(User.username == lector_username, User.rol == "lector")
            )
            lector = lector_result.scalar_one_or_none()
            if lector:
                lector_id = lector.id
            else:
                errores.append({"fila": i, "codigo": codigo, "error": f"Lector '{lector_username}' no encontrado"})
                continue

        try:
            latitud = float(row["latitud"]) if row.get("latitud", "").strip() else None
        except (ValueError, TypeError):
            latitud = None
        try:
            longitud = float(row["longitud"]) if row.get("longitud", "").strip() else None
        except (ValueError, TypeError):
            longitud = None

        meter = Meter(
            codigo_medidor=codigo,
            niud=row.get("niud", "").strip() or None,
            direccion=row.get("direccion", "").strip() or None,
            vereda=row.get("vereda", "").strip() or None,
            ruta=row.get("ruta", "").strip() or None,
            latitud=latitud,
            longitud=longitud,
            tipo=row.get("tipo", "").strip() or None,
            estado=row.get("estado", "activo").strip(),
            lector_id=lector_id,
        )
        session.add(meter)
        creados += 1

    await session.flush()

    await log_action(
        session, accion="meter.bulk_create", entidad_tipo="meter",
        entidad_id=current_user.id, actor_id=str(current_user.id),
        detalle={"creados": creados, "errores": len(errores)},
    )

    return {
        "creados": creados,
        "errores": errores,
        "total_procesadas": len(rows),
    }


@router.post("/cargar-lecturas")
async def cargar_lecturas(
    session: SessionDep,
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles("admin", "supervisor")),
):
    from datetime import date as date_type

    filename = file.filename or ""
    content = await file.read()

    rows: list[dict[str, str]] = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="El archivo XLSX no tiene hojas")
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o XLSX")

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    creados = 0
    errores: list[dict] = []

    for i, row in enumerate(rows, start=2):
        meter_codigo = row.get("meter_codigo", "").strip()
        if not meter_codigo:
            errores.append({"fila": i, "error": "meter_codigo es requerido"})
            continue

        meter_result = await session.execute(
            select(Meter).where(Meter.codigo_medidor == meter_codigo)
        )
        meter = meter_result.scalar_one_or_none()
        if not meter:
            errores.append({"fila": i, "codigo": meter_codigo, "error": "Medidor no encontrado"})
            continue

        lector_id: UUID | None = current_user.id
        lector_username = row.get("lector_username", "").strip()
        if lector_username:
            lector_result = await session.execute(
                select(User).where(User.username == lector_username, User.rol == "lector")
            )
            lector = lector_result.scalar_one_or_none()
            if lector:
                lector_id = lector.id
            else:
                errores.append({"fila": i, "codigo": meter_codigo, "error": f"Lector '{lector_username}' no encontrado"})
                continue

        try:
            lectura_actual = float(row.get("lectura_actual", "").strip() or "0")
        except ValueError:
            errores.append({"fila": i, "codigo": meter_codigo, "error": "lectura_actual inválida"})
            continue

        fecha_str = row.get("fecha_lectura", "").strip()
        try:
            fecha_lectura = date_type.fromisoformat(fecha_str)
        except ValueError:
            errores.append({"fila": i, "codigo": meter_codigo, "error": f"fecha_lectura inválida: '{fecha_str}' (use YYYY-MM-DD)"})
            continue

        last_result = await session.execute(
            select(Reading)
            .where(Reading.meter_id == meter.id)
            .order_by(Reading.fecha_lectura.desc())
            .limit(1)
        )
        last_reading = last_result.scalar_one_or_none()
        lectura_anterior = last_reading.lectura_actual if last_reading else None
        consumo = None
        if lectura_anterior is not None and lectura_actual >= lectura_anterior:
            consumo = lectura_actual - lectura_anterior

        reading = Reading(
            meter_id=meter.id,
            lector_id=lector_id,
            lectura_anterior=lectura_anterior,
            lectura_actual=lectura_actual,
            consumo=consumo,
            fecha_lectura=fecha_lectura,
            metodo_captura=row.get("metodo_captura", "manual").strip(),
            estado_sync="synced",
            estado_validacion="validada",
        )
        session.add(reading)
        creados += 1

    await session.flush()

    await log_action(
        session, accion="reading.bulk_create", entidad_tipo="reading",
        entidad_id=current_user.id, actor_id=str(current_user.id),
        detalle={"creados": creados, "errores": len(errores)},
    )

    return {
        "creados": creados,
        "errores": errores,
        "total_procesadas": len(rows),
    }
